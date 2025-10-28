"""CLI tool to convert Excel workbooks into TSV files per worksheet."""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path
from typing import Iterable, List, Sequence

from openpyxl import load_workbook


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Convert each worksheet in an Excel workbook into a TSV file.",
    )
    parser.add_argument(
        "-i", "--input", required=True, help="Path to the input Excel workbook"
    )
    parser.add_argument(
        "-o", "--output", help="Directory where TSV files will be written"
    )
    parser.add_argument(
        "-c",
        "--stdout",
        action="store_true",
        help="Print TSV content to stdout instead of writing files",
    )
    return parser


def sanitize_sheet_name(name: str) -> str:
    """Create a filesystem-friendly representation for a worksheet name."""
    cleaned = name.strip()
    cleaned = re.sub(r"[^\w.-]+", "_", cleaned)
    return cleaned or "sheet"


def format_row(values: Sequence[object]) -> List[str]:
    """Convert cell values to strings while trimming trailing empty cells."""
    row = ["" if value is None else str(value) for value in values]
    while row and row[-1] == "":
        row.pop()
    return row


def iter_worksheets(workbook) -> Iterable:
    """Yield worksheets preserving workbook order (extracted for testing)."""
    return workbook.worksheets


def main(argv: List[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    input_path = Path(args.input)
    if not input_path.exists():
        parser.error(f"Input workbook '{input_path}' does not exist.")

    if not args.stdout and not args.output:
        parser.error("Either provide --output or use --stdout to print results.")

    output_dir = Path(args.output) if args.output else None
    if output_dir:
        output_dir.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(filename=input_path, data_only=True)
    sheet_count = 0

    for worksheet in iter_worksheets(workbook):
        sheet_count += 1
        sheet_name = sanitize_sheet_name(worksheet.title)
        if args.stdout:
            print(f"# Worksheet: {worksheet.title}")
            writer = csv.writer(sys.stdout, delimiter="\t", lineterminator="\n")
            row_count = 0
            max_cols = 0
            for values in worksheet.iter_rows(values_only=True):
                row = format_row(values)
                writer.writerow(row)
                row_count += 1
                if row:
                    max_cols = max(max_cols, len(row))
            print(f"# End worksheet ({row_count} rows, {max_cols} columns)\n")
        else:
            assert output_dir is not None
            output_file = output_dir / f"{sheet_name}.tsv"
            with output_file.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.writer(handle, delimiter="\t", lineterminator="\n")
                row_count = 0
                max_cols = 0
                for values in worksheet.iter_rows(values_only=True):
                    row = format_row(values)
                    writer.writerow(row)
                    row_count += 1
                    if row:
                        max_cols = max(max_cols, len(row))
            print(f"Created {output_file} ({row_count} rows, {max_cols} columns)")

    if sheet_count == 0:
        print(f"No worksheets found in '{input_path}'.")
    else:
        print(f"Processed {sheet_count} worksheet(s) from '{input_path}'.")

    return 0


if __name__ == "__main__":  # pragma: no cover
    sys.exit(main())
