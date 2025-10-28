"""Tests for the gan-init command line interface."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from gan_nomenclature import init_cli


def get_sheet_rows(path: Path) -> list[list[str | None]]:
    workbook = load_workbook(path)
    worksheet = workbook.active
    return [[cell.value for cell in row] for row in worksheet.iter_rows(values_only=False)]


def test_gan_init_creates_empty_templates(tmp_path):
    """Templates should contain only headers unless --example is provided."""
    outdir = tmp_path / "templates"
    exit_code = init_cli.main(["-o", str(outdir)])

    assert exit_code == 0

    for filename in (init_cli.PRIMARY_FILENAME, init_cli.SECONDARY_FILENAME):
        path = outdir / filename
        assert path.exists()
        rows = get_sheet_rows(path)
        assert len(rows) == 1
        assert rows[0] == list(init_cli.COLUMNS)


def test_gan_init_populates_examples(tmp_path, capsys):
    """--example should add predefined rows (and emit a helpful message)."""
    outdir = tmp_path / "examples"
    exit_code = init_cli.main(["-o", str(outdir), "--example"])

    assert exit_code == 0

    for filename in (init_cli.PRIMARY_FILENAME, init_cli.SECONDARY_FILENAME):
        path = outdir / filename
        rows = get_sheet_rows(path)
        assert len(rows) == 1 + len(init_cli.EXAMPLE_ROWS)
        for expected, actual in zip(init_cli.EXAMPLE_ROWS, rows[1:]):
            assert actual == [expected.get(column, "") for column in init_cli.COLUMNS]

    output = capsys.readouterr().out
    assert "Templates populated with example rows." in output
